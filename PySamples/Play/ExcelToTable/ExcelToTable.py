from openpyxl import load_workbook

from pyrx import Ap, Db


def PyRxCmd_pydoit():
    try:
        table = Db.Table()
        table.setSize(22, 8)
        table.generateLayout()
        table.setDatabaseDefaults()

        wb = load_workbook("sample.xlsx")
        sheet = wb.active
        for icol in range(7):
            for irow in range(20):
                val = sheet.cell(row=irow + 1, column=icol + 1)
                table.setTextString(irow, icol, "{}".format(val.value))

        db = Ap.Application().docManager().curDocument().database()
        model = Db.BlockTableRecord(db.modelSpaceId(), Db.OpenMode.kForWrite)
        model.appendAcDbEntity(table)

    except Exception as err:
        print("Error: {}".format(err))
