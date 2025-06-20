import traceback

import openpyxl as xl
import pandas as pd

from pyrx import Db, Ed


# create from excel
def PyRxCmd_doit():
    try:
        wb = xl.load_workbook("E:\\ItsAlive.xlsx", read_only=True)
        ws = wb.active

        table = Db.Table()
        table.setDatabaseDefaults()

        # add one for the title and header
        table.setSize(ws.max_row + 2, ws.max_column)
        table.generateLayout()

        # note openpyxl cell indexing is not zero based ..?..
        for row in ws.iter_rows():
            for cell in row:
                table.setTextString(cell.row + 1, cell.column - 1, "{}".format(cell.value))

        db = Db.curDb()
        model = Db.BlockTableRecord(db.modelSpaceId(), Db.OpenMode.kForWrite)
        model.appendAcDbEntity(table)

    except Exception as err:
        traceback.print_exception(err)


# create from pandas
def PyRxCmd_doit2() -> None:
    try:
        db = Db.curDb()

        data = {"calories": [420, 380, 390], "duration": [50, 40, 45]}

        df = pd.DataFrame(data)

        table = Db.Table()
        table.setDatabaseDefaults(db)

        # add one for the title and header
        table.setSize(df.shape[0] + 2, df.shape[1])

        # title
        table.setTextString(0, 0, "I Like to Move It, Move It")

        headers = df.columns.values.tolist()
        datas = df.values.tolist()

        for col, value in enumerate(headers):
            table.setTextString(1, col, "{}".format(value))

        for row, data in enumerate(datas):
            for col, value in enumerate(data):
                table.setTextString(row + 2, col, "{}".format(value))

        model = Db.BlockTableRecord(db.modelSpaceId(), Db.OpenMode.kForWrite)
        model.appendAcDbEntity(table)

    except Exception:
        print(traceback.format_exc())


# simple iteration
def PyRxCmd_doit3() -> None:
    try:
        es = Ed.Editor.entSel("\nSelect a table: ", Db.Table.desc())
        if es[0] != Ed.PromptStatus.eNormal:
            return

        table = Db.Table(es[1])
        for cell in table.cells(Db.CellRange(1, 0, 3, 3)):
            print(table.textString(cell))

    except Exception as err:
        traceback.print_exception(err)


# simple iteration cellValues preserves the type. int, float, string
def PyRxCmd_doit4() -> None:
    try:
        es = Ed.Editor.entSel("\nSelect a table: ", Db.Table.desc())
        if es[0] != Ed.PromptStatus.eNormal:
            return

        opts = Db.TableIteratorOption.kTableIteratorSkipMerged

        table = Db.Table(es[1])
        for cellValue in table.cellValues(opts):
            print(cellValue)

    except Exception as err:
        traceback.print_exception(err)


# simple iteration cellStrValues, returns list of tuples (nRow, nCol, String)
def PyRxCmd_doit5() -> None:
    try:
        es = Ed.Editor.entSel("\nSelect: ", Db.Table.desc())
        table = Db.Table(es[1])

        opts = Db.TableIteratorOption.kTableIteratorSkipMerged
        for item in table.cellStrValues(opts):
            if item[2] != "":
                print(item)

    except Exception:
        traceback.print_exc()


def PyRxCmd_setMargins1():
    try:
        db = Db.curDb()
        table = Db.Table()
        table.setDatabaseDefaults(db)
        table.setTableStyle(db.tablestyle())
        table.setVertCellMargin(0.07)
        table.setHorzCellMargin(0.07)
        table.setSize(5, 5)
        table.generateLayout()
        db.addToModelspace(table)

        print(table.vertCellMargin())
        print(table.horzCellMargin())

    except Exception as err:
        traceback.print_exception(err)


def PyRxCmd_setMargins2():
    try:
        db = Db.curDb()
        table = Db.Table()
        table.setDatabaseDefaults(db)
        table.setTableStyle(db.tablestyle())
        table.setSize(5, 5)
        table.generateLayout()
        db.addToModelspace(table)

        v = Db.CellMargin.kCellMarginLeft | Db.CellMargin.kCellMarginRight
        h = Db.CellMargin.kCellMarginTop | Db.CellMargin.kCellMarginBottom

        table.setMargin(-1, -1, Db.CellMargin(v), 0.08)
        table.setMargin(-1, -1, Db.CellMargin(h), 0.08)

    except Exception as err:
        traceback.print_exception(err)
