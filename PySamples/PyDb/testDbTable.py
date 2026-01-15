"""
PyRx Command Module for Table Creation and Manipulation from Excel/Pandas Data.

This module provides multiple commands to create AutoCAD tables from various data sources,
including Excel files and pandas DataFrames, along with functions to iterate through
table cells and manipulate table properties like margins.

Commands:
    doit   - Creates a table from an Excel file (E:\\ItsAlive.xlsx)
    doit2  - Creates a table from pandas DataFrame with sample data
    doit3  - Iterates through selected table cells and prints their text values
    doit4  - Iterates through selected table cell values preserving data types
    doit5  - Iterates through selected table cell string values

Functions:
    doit() -> None
        Creates a table from Excel file data
        
    doit2() -> None  
        Creates a table from pandas DataFrame with sample data
        
    PyRxCmd_doit3() -> None
        Iterates through table cells and prints text content
        
    PyRxCmd_doit4() -> None
        Iterates through table cell values preserving data types (int, float, string)
        
    PyRxCmd_doit5() -> None
        Iterates through table cell string values and filters empty strings
        
    PyRxCmd_setMargins1() -> None
        Creates a table with custom cell margins using vertical/horizontal settings
        
    PyRxCmd_setMargins2() -> None
        Creates a table with custom cell margins using specific margin types

Example usage:
    Run any of the commands in AutoCAD/BricsCAD to create or manipulate tables.
"""

import traceback

import openpyxl as xl
import pandas as pd

from pyrx import Ap, Db, Ed


# create from excel
@Ap.Command()
def doit() -> None:
    """
    Create an AutoCAD table from data in an Excel file.

    This function reads data from E:\\ItsAlive.xlsx and creates a corresponding 
    AutoCAD table with the same structure. The table includes all rows and columns 
    from the Excel sheet, with proper positioning of data cells.

    Features:
    - Loads Excel workbook in read-only mode for performance
    - Creates AutoCAD table with size matching Excel dimensions + 2 extra rows (for title/header)
    - Transfers all cell values from Excel to AutoCAD table
    - Positions the table at default location in model space

    Example usage:
        Run 'doit' command in AutoCAD to create table from E:\\ItsAlive.xlsx
        
    Note: 
        OpenPyXL uses 1-based indexing for rows and columns, while AutoCAD uses 0-based.
        The function adjusts for this difference when mapping cells.
    """
    try:
        wb = xl.load_workbook("E:\\ItsAlive.xlsx", read_only=True)
        ws = wb.active

        table = Db.Table()
        table.setDatabaseDefaults()

        # add one for the title and header
        table.setSize(ws.max_row + 2, ws.max_column)
        table.generateLayout()

        # note openpyxl cell indexing is not zero based ..?..
        for row in ws.iter_rows():
            for cell in row:
                table.setTextString(
                    cell.row + 1, cell.column - 1, "{}".format(cell.value)
                )

        db = Db.curDb()
        model = Db.BlockTableRecord(db.modelSpaceId(), Db.OpenMode.kForWrite)
        model.appendAcDbEntity(table)

    except Exception as err:
        traceback.print_exception(err)


# create from pandas
@Ap.Command()
def doit2() -> None:
    """
    Create an AutoCAD table from a pandas DataFrame.

    This function creates a sample table with predefined data using pandas,
    then transfers that data to an AutoCAD table entity. The resulting table 
    includes column headers, data rows, and a title row.

    Features:
    - Creates sample data: calories [420, 380, 390] and duration [50, 40, 45]
    - Uses pandas DataFrame for structured data handling
    - Generates AutoCAD table with proper sizing based on DataFrame dimensions  
    - Adds title row (row 0) with custom text
    - Populates header row (row 1) with column names
    - Fills data rows (starting from row 2) with values from DataFrame

    Example usage:
        Run 'doit2' command in AutoCAD to create table from pandas sample data
        
    Data structure created:
        Row 0: "I Like to Move It, Move It" (title)
        Row 1: "calories", "duration" (headers)  
        Rows 2-4: [420, 50], [380, 40], [390, 45] (data)
    """
    try:
        db = Db.curDb()

        data = {"calories": [420, 380, 390], "duration": [50, 40, 45]}

        df = pd.DataFrame(data)

        table = Db.Table()
        table.setDatabaseDefaults(db)

        # add one for the title and header
        table.setSize(df.shape[0] + 2, df.shape[1])

        # title
        table.setTextString(0, 0, "I Like to Move It, Move It")

        headers = df.columns.values.tolist()
        datas = df.values.tolist()

        for col, value in enumerate(headers):
            table.setTextString(1, col, "{}".format(value))

        for row, data in enumerate(datas):
            for col, value in enumerate(data):
                table.setTextString(row + 2, col, "{}".format(value))

        model = Db.BlockTableRecord(db.modelSpaceId(), Db.OpenMode.kForWrite)
        model.appendAcDbEntity(table)

    except Exception:
        print(traceback.format_exc())


# simple iteration
@Ap.Command()
def doit3() -> None:
    """
    Iterate through a selected table's cells and print their text values.

    This function prompts the user to select an existing AutoCAD table, then
    iterates through a specified range of cells (rows 1-3, columns 0-3) 
    and prints each cell's text content to the console.

    Features:
    - Prompts user for table selection via entity pick
    - Iterates through predefined cell range using Db.CellRange
    - Prints each cell's text value using table.textString()
    - Handles prompt status checking for robust operation

    Example usage:
        Run 'doit3' command in AutoCAD, then select a table to see its contents
        
    Cell range processed: rows 1-3, columns 0-3
    """
    try:
        es = Ed.Editor.entSel("\nSelect a table: ", Db.Table.desc())
        if es[0] != Ed.PromptStatus.eNormal:
            return

        table = Db.Table(es[1])
        for cell in table.cells(Db.CellRange(1, 0, 3, 3)):
            print(table.textString(cell))

    except Exception as err:
        traceback.print_exception(err)


# simple iteration cellValues preserves the type. int, float, string
@Ap.Command()
def doit4() -> None:
    """
    Iterate through a selected table's cells and print their values preserving data types.

    This function prompts the user to select an existing AutoCAD table, then 
    iterates through all cells (excluding merged cells) and prints each cell's 
    value with its preserved original data type (int, float, string).

    Features:
    - Prompts user for table selection via entity pick
    - Uses Db.TableIteratorOption.kTableIteratorSkipMerged to avoid duplicate processing
    - Iterates through all valid cells in the table
    - Prints cell values while preserving their original data types
    - Handles prompt status checking

    Example usage:
        Run 'doit4' command in AutoCAD, then select a table to see its typed values
        
    Note: 
        Unlike textString(), this method preserves numeric types (int/float) rather than converting everything to string.
    """
    try:
        es = Ed.Editor.entSel("\nSelect a table: ", Db.Table.desc())
        if es[0] != Ed.PromptStatus.eNormal:
            return

        opts = Db.TableIteratorOption.kTableIteratorSkipMerged

        table = Db.Table(es[1])
        for cellValue in table.cellValues(opts):
            print(cellValue)

    except Exception as err:
        traceback.print_exception(err)


# simple iteration cellStrValues, returns list of tuples (nRow, nCol, String)
@Ap.Command()
def doit5() -> None:
    """
    Iterate through a selected table's cells and print string representations.

    This function prompts the user to select an AutoCAD table, then iterates
    through all cells (excluding merged cells) and prints each cell's 
    position and string representation as tuples of (row, column, text).

    Features:
    - Prompts user for table selection via entity pick  
    - Uses Db.TableIteratorOption.kTableIteratorSkipMerged to skip merged cells
    - Returns data as tuples: (row_index, column_index, cell_text_string)
    - Filters out empty string values and prints only non-empty cells
    - Handles prompt status checking

    Example usage:
        Run 'doit5' command in AutoCAD, then select a table to see formatted cell info
        
    Output format:
        (row_number, column_number, "cell_content")
        
    Note: 
        This method is useful for processing table data where you need the exact
        position and string representation of each cell.
    """
    try:
        es = Ed.Editor.entSel("\nSelect: ", Db.Table.desc())
        table = Db.Table(es[1])

        opts = Db.TableIteratorOption.kTableIteratorSkipMerged
        for item in table.cellStrValues(opts):
            if item[2] != "":
                print(item)

    except Exception:
        traceback.print_exc()


@Ap.Command()
def setMargins1():
    """
    Create a table with custom cell margins using vertical/horizontal properties.

    This function demonstrates setting uniform cell margins across all cells
    in an AutoCAD table using the vertCellMargin and horzCellMargin properties.
    
    Features:
    - Creates new AutoCAD table with default settings
    - Sets table style from current database
    - Configures both vertical and horizontal cell margins to 0.07 units
    - Sizes table to 5 rows x 5 columns  
    - Generates layout and adds to model space
    - Prints the margin values after creation for verification

    Example usage:
        Run 'setMargins1' command in AutoCAD to create a table with custom spacing
        
    Margin settings: 
        Vertical margins = 0.07 units
        Horizontal margins = 0.07 units
    """
    try:
        db = Db.curDb()
        table = Db.Table()
        table.setDatabaseDefaults(db)
        table.setTableStyle(db.tablestyle())
        table.setVertCellMargin(0.07)
        table.setHorzCellMargin(0.07)
        table.setSize(5, 5)
        table.generateLayout()
        db.addToModelspace(table)

        print(table.vertCellMargin())
        print(table.horzCellMargin())

    except Exception as err:
        traceback.print_exception(err)


@Ap.Command()
def setMargins2():
    """
    Create a table with custom cell margins using specific margin types.

    This function demonstrates setting different margin values for specific 
    sides of cells in an AutoCAD table by specifying individual margin types.
    
    Features:
    - Creates new AutoCAD table with default settings
    - Sets table style from current database
    - Sizes table to 5 rows x 5 columns
    - Generates layout and adds to model space
    - Sets left/right margins for all cells using Db.CellMargin.kCellMarginLeft | kCellMarginRight
    - Sets top/bottom margins for all cells using Db.CellMargin.kCellMarginTop | kCellMarginBottom  
    - Applies margin value of 0.08 units to specified sides

    Example usage:
        Run 'setMargins2' command in AutoCAD to create a table with custom side-specific spacing
        
    Margin settings: 
        All cells: Left/Right = 0.08 units, Top/Bottom = 0.08 units
    """
    try:
        db = Db.curDb()
        table = Db.Table()
        table.setDatabaseDefaults(db)
        table.setTableStyle(db.tablestyle())
        table.setSize(5, 5)
        table.generateLayout()
        db.addToModelspace(table)

        v = Db.CellMargin.kCellMarginLeft | Db.CellMargin.kCellMarginRight
        h = Db.CellMargin.kCellMarginTop | Db.CellMargin.kCellMarginBottom

        table.setMargin(-1, -1, Db.CellMargin(v), 0.08)
        table.setMargin(-1, -1, Db.CellMargin(h), 0.08)

    except Exception as err:
        traceback.print_exception(err)
